"""
Aggregate vLLM benchmark CSVs into per-model xlsx reports and bar charts.

Run:
    python perf_report.py --config config.json

The config JSON must contain:
    source_dir       Directory whose subdirectories are app names (one app per subdir).
    output_dir       Where the run directory will be created.
    baseline_app     Name of the baseline app (matches one of the discovered subdirs,
                     case-insensitively).
    models           List of model labels, e.g. ["mixtral", "qwen", "llama"]. Each
                     CSV's `model_id` column is mapped to one of these by
                     case-insensitive substring match.
    metrics          List of metric column names to aggregate. Defaults to
                     ["total_token_throughput", "mean_ttft_ms", "mean_itl_ms", "p99_itl_ms"].
    warmup_rows      Number of leading rows per config to drop. Default: 5.
"""

from __future__ import annotations

import argparse
import json
import logging
import math
import shutil
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterable

import matplotlib.pyplot as plt
import matplotlib.patheffects as pe
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.error_bar import ErrorBars
from openpyxl.chart.data_source import NumDataSource, NumRef
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

DEFAULT_METRICS = ["total_token_throughput", "mean_ttft_ms", "mean_itl_ms", "p99_itl_ms"]
DEFAULT_WARMUP = 5
CONFIG_COLS = ["max_num_seqs"]
MODEL_ID_COL = "model_id"
COMBINED_LABEL = "all_apps_combined"
ACROSS_MODELS_LABEL = "across_models"
DEFAULT_HATCH_PATTERNS = ["", "//", "\\\\", "xx", "..", "++", "oo", "--"]
DEFAULT_SPECIAL_COLORS = {
    COMBINED_LABEL: "#d62728",
    ACROSS_MODELS_LABEL: "#17becf",
}
CHART_YMAX = 10.0
CHART_OVERFLOW_EPS = 0.15

# Bootstrap CI parameters for app/baseline ratio comparison.
# - Latency metrics (lower-is-better): use arithmetic mean. Slowdown = updated/baseline.
# - Throughput metrics (higher-is-better): use harmonic mean (correct for rates).
#   Speedup = updated/baseline; reported as % change with the same convention as
#   before (positive = regression).
# CI is computed by independently resampling the two arms with replacement and
# taking the percentile of the resulting ratio distribution.
N_BOOTSTRAP = 2000
CI_LEVEL = 0.95
_BOOT_RNG = np.random.default_rng(0)  # deterministic CIs across runs

log = logging.getLogger("perf_report")


# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------


@dataclass
class Config:
    source_dir: Path
    output_dir: Path
    baseline_app: str
    models: list[str]
    metrics: list[str]
    warmup_rows: int
    chart_presentation: "ChartPresentation"

    @classmethod
    def from_json(cls, path: Path) -> "Config":
        with path.open() as f:
            raw = json.load(f)
        missing = {"source_dir", "output_dir", "baseline_app", "models"} - raw.keys()
        if missing:
            raise ValueError(f"Config missing required keys: {sorted(missing)}")
        return cls(
            source_dir=Path(raw["source_dir"]),
            output_dir=Path(raw["output_dir"]),
            baseline_app=raw["baseline_app"],
            models=list(raw["models"]),
            metrics=list(raw.get("metrics", DEFAULT_METRICS)),
            warmup_rows=int(raw.get("warmup_rows", DEFAULT_WARMUP)),
            chart_presentation=ChartPresentation.from_raw(raw.get("chart_presentation")),
        )


@dataclass
class ChartPresentation:
    hidden_apps: set[str] = field(default_factory=lambda: {"vanilla_run", "app_5"})
    app_order: list[str] = field(default_factory=list)
    label_map: dict[str, str] = field(default_factory=dict)
    series_name_map: dict[str, str] = field(default_factory=dict)
    font_sizes: dict[str, dict[str, float]] = field(default_factory=dict)
    font_weights: dict[str, dict[str, str]] = field(default_factory=dict)
    colors: dict[str, str] = field(default_factory=dict)
    hatches: dict[str, object] = field(default_factory=dict)
    metric_axis_labels: dict[str, str] = field(default_factory=dict)
    text: dict[str, str] = field(default_factory=dict)

    @classmethod
    def default_raw(cls) -> dict[str, object]:
        return {
            "hidden_apps": ["vanilla_run", "app_5"],
            "app_order": [],
            "label_map": {
                "vanilla_run": "Vanilla Run",
                "app_1": "App#1",
                "app_2": "App#2",
                "app_3": "App#3",
                "app_4": "App#4",
                "app_5": "App#5",
                COMBINED_LABEL: "All Apps Combined",
                ACROSS_MODELS_LABEL: "Across-models",
            },
            "series_name_map": {
                "app_1": "App#1",
                "app_2": "App#2",
                "app_3": "App#3",
                "app_4": "App#4",
                "app_5": "App#5",
                COMBINED_LABEL: "All Apps Combined",
                ACROSS_MODELS_LABEL: "Across-models",
            },
            "font_sizes": {
                "default": {
                    "base": 11,
                    "title": 20,
                    "axes_label": 20,
                    "xtick": 16,
                    "ytick": 14,
                    "annotation": 12,
                    "legend": 20,
                },
                "label_overrides": {},
            },
            "font_weights": {
                "default": {},
                "label_overrides": {
                    COMBINED_LABEL: {
                        "legend": "semibold",
                    },
                    ACROSS_MODELS_LABEL: {
                        "xtick": "semibold",
                        "legend": "semibold",
                    },
                },
            },
            "colors": dict(DEFAULT_SPECIAL_COLORS),
            "hatches": {
                "default_patterns": DEFAULT_HATCH_PATTERNS,
                "series_map": {
                    "vanilla_run": "",
                    "app_1": "//",
                    "app_2": "\\\\",
                    "app_3": "xx",
                    "app_4": "..",
                    "app_5": "++",
                    COMBINED_LABEL: "--",
                    ACROSS_MODELS_LABEL: "oo",
                },
            },
            "metric_axis_labels": {
                "mean_itl_ms": "Mean ITL overhead (%, ↓)",
                "p99_itl_ms": "P99 ITL overhead (%, ↓)",
                "mean_ttft_ms": "Mean TTFT overhead (%, ↓)",
                "total_token_throughput": "Throughput loss (%, ↓)",
            },
            "text": {
                "y_axis_label_overhead": "% Overhead",
                "y_axis_label_throughput": "% Throughput",
            },
        }

    @classmethod
    def from_raw(cls, raw: dict[str, object] | None) -> "ChartPresentation":
        merged = cls.default_raw()
        if raw:
            for key, value in raw.items():
                if key in {"label_map", "series_name_map", "colors", "metric_axis_labels", "text"} and isinstance(value, dict):
                    merged[key].update(value)
                elif key == "font_sizes" and isinstance(value, dict):
                    merged_defaults = merged["font_sizes"].get("default", {})
                    merged_overrides = merged["font_sizes"].get("label_overrides", {})
                    merged["font_sizes"] = {
                        "default": {**merged_defaults, **value.get("default", {})},
                        "label_overrides": {
                            **merged_overrides,
                            **value.get("label_overrides", {}),
                        },
                    }
                elif key == "font_weights" and isinstance(value, dict):
                    merged_defaults = merged["font_weights"].get("default", {})
                    merged_overrides = merged["font_weights"].get("label_overrides", {})
                    merged["font_weights"] = {
                        "default": {
                            **merged_defaults,
                            **{str(k): str(v) for k, v in value.get("default", {}).items()},
                        },
                        "label_overrides": {
                            **merged_overrides,
                            **{
                                str(k): {str(rk): str(rv) for rk, rv in v.items()}
                                for k, v in value.get("label_overrides", {}).items()
                                if isinstance(v, dict)
                            },
                        },
                    }
                elif key == "hatches" and isinstance(value, dict):
                    merged["hatches"] = {
                        "default_patterns": list(
                            value.get(
                                "default_patterns",
                                merged["hatches"].get("default_patterns", DEFAULT_HATCH_PATTERNS),
                            )
                        ),
                        "series_map": {
                            **merged["hatches"].get("series_map", {}),
                            **value.get("series_map", {}),
                        },
                    }
                else:
                    merged[key] = value
        return cls(
            hidden_apps={str(v) for v in merged.get("hidden_apps", [])},
            app_order=[str(v) for v in merged.get("app_order", [])],
            label_map={str(k): str(v) for k, v in merged.get("label_map", {}).items()},
            series_name_map={str(k): str(v) for k, v in merged.get("series_name_map", {}).items()},
            font_sizes={
                "default": {
                    str(k): float(v)
                    for k, v in merged.get("font_sizes", {}).get("default", {}).items()
                },
                "label_overrides": {
                    str(k): {str(rk): float(rv) for rk, rv in v.items()}
                    for k, v in merged.get("font_sizes", {}).get("label_overrides", {}).items()
                    if isinstance(v, dict)
                },
            },
            font_weights={
                "default": {
                    str(k): str(v)
                    for k, v in merged.get("font_weights", {}).get("default", {}).items()
                },
                "label_overrides": {
                    str(k): {str(rk): str(rv) for rk, rv in v.items()}
                    for k, v in merged.get("font_weights", {}).get("label_overrides", {}).items()
                    if isinstance(v, dict)
                },
            },
            colors={str(k): str(v) for k, v in merged.get("colors", {}).items()},
            hatches={
                "default_patterns": [str(v) for v in merged.get("hatches", {}).get("default_patterns", DEFAULT_HATCH_PATTERNS)],
                "series_map": {
                    str(k): str(v)
                    for k, v in merged.get("hatches", {}).get("series_map", {}).items()
                },
            },
            metric_axis_labels={
                str(k): str(v)
                for k, v in merged.get("metric_axis_labels", {}).items()
            },
            text={str(k): str(v) for k, v in merged.get("text", {}).items()},
        )

    def hide_app(self, app: str) -> bool:
        return app in self.hidden_apps

    def axis_text(self, key: str, default: str) -> str:
        return self.text.get(key, default)

    def display_label(self, key: str) -> str:
        return self.label_map.get(key, key)

    def series_display_name(self, key: str) -> str:
        return self.series_name_map.get(key, self.display_label(key))

    @staticmethod
    def _normalize_label_key(value: str) -> str:
        return value.replace("\n", " ").replace("-", "_").replace(" ", "_").lower()

    def _override_candidates(self, label_key: str) -> list[str]:
        candidate_keys = [label_key]

        display_label = self.display_label(label_key)
        if display_label not in candidate_keys:
            candidate_keys.append(display_label)

        normalized_target = self._normalize_label_key(label_key)
        normalized_display = self._normalize_label_key(display_label)
        return candidate_keys + [normalized_target, normalized_display]

    def _matching_override_keys(self, overrides: dict[str, dict[str, object]], label_key: str) -> list[str]:
        candidate_keys = self._override_candidates(label_key)
        normalized_candidates = {self._normalize_label_key(key) for key in candidate_keys}
        matches: list[str] = []
        seen: set[str] = set()
        for candidate in candidate_keys:
            if candidate in overrides and candidate not in seen:
                matches.append(candidate)
                seen.add(candidate)
        for override_key in overrides:
            if override_key in seen:
                continue
            if self._normalize_label_key(override_key) in normalized_candidates:
                matches.append(override_key)
                seen.add(override_key)
        return matches

    def font_size(self, role: str, label_key: str | None = None) -> float:
        default = self.font_sizes.get("default", {}).get(role)
        if default is None:
            default = self.font_sizes.get("default", {}).get("base", 11.0)
        if label_key is None:
            return float(default)
        overrides = self.font_sizes.get("label_overrides", {})
        for candidate in self._matching_override_keys(overrides, label_key):
            value = overrides.get(candidate, {}).get(role)
            if value is not None:
                return float(value)
        return float(default)

    def font_weight(self, role: str, label_key: str | None = None) -> str:
        default = str(self.font_weights.get("default", {}).get(role, "normal"))
        if label_key is None:
            return default
        overrides = self.font_weights.get("label_overrides", {})
        for candidate in self._matching_override_keys(overrides, label_key):
            value = overrides.get(candidate, {}).get(role)
            if value is not None:
                return str(value)
        return default

    def color(self, key: str, default: str) -> str:
        return self.colors.get(key, default)

    def hatch_patterns(self) -> list[str]:
        patterns = self.hatches.get("default_patterns", DEFAULT_HATCH_PATTERNS)
        return patterns if patterns else DEFAULT_HATCH_PATTERNS

    def explicit_hatch(self, key: str) -> str | None:
        return self.hatches.get("series_map", {}).get(key)


def _metric_axis_label(metric: str, presentation: ChartPresentation) -> str:
    axis_label = presentation.metric_axis_labels.get(metric)
    if axis_label is not None:
        return axis_label
    text_key = "y_axis_label_throughput" if _is_throughput_metric(metric) else "y_axis_label_overhead"
    return presentation.text[text_key]


# ---------------------------------------------------------------------------
# Discovery & loading
# ---------------------------------------------------------------------------


def discover_apps(source_dir: Path) -> dict[str, list[tuple[Path, str]]]:
    """Return {raw_subdir_name: [(csv_path, raw_subdir_name), ...]}.

    Returned as raw subdirectory names; app inference happens later because it
    needs the configured model list.
    """
    apps: dict[str, list[tuple[Path, str]]] = {}
    for sub in sorted(p for p in source_dir.iterdir() if p.is_dir()):
        csvs = sorted(sub.glob("summary_*.csv"))
        if csvs:
            apps[sub.name] = [(c, sub.name) for c in csvs]
        else:
            log.debug("No summary_*.csv files in %s, skipping", sub)
    if not apps:
        raise RuntimeError(f"No subdirectories with summary_*.csv found under {source_dir}")
    return apps


def infer_app_name(dir_name: str, models: list[str]) -> str | None:
    """Strip the timestamp/gpu/model prefix from a run-id-style directory name.

    Rule: find the configured model substring (case-insensitive), and take
    everything after it as the app name (with surrounding underscores stripped).
    Returns None if no model substring is present.
    """
    low = dir_name.lower()
    # Prefer the longest model match in case of overlap (e.g. "llama2" vs "llama")
    candidates = sorted(
        ((m, low.rfind(m.lower())) for m in models if m.lower() in low),
        key=lambda mp: (mp[1], len(mp[0])),
        reverse=True,
    )
    if not candidates:
        return None
    model, idx = candidates[0]
    suffix = dir_name[idx + len(model):]
    app = suffix.strip("_-. ")
    return app or None


def resolve_baseline(app_names: set[str], baseline_app: str) -> str:
    """Case-insensitive exact match against the inferred app names."""
    matches = [name for name in app_names if name.lower() == baseline_app.lower()]
    if not matches:
        raise RuntimeError(
            f"Baseline app {baseline_app!r} not found among inferred apps: "
            f"{sorted(app_names)}"
        )
    return matches[0]


def map_model(model_id: str, models: list[str]) -> str | None:
    """Map a CSV's model_id column to one of the configured model labels."""
    model_id_low = model_id.lower()
    matches = [m for m in models if m.lower() in model_id_low]
    if len(matches) == 1:
        return matches[0]
    if len(matches) > 1:
        log.warning(
            "model_id %r matched multiple configured models %s; skipping",
            model_id, matches,
        )
        return None
    return None


def load_csv_post_warmup(path: Path, warmup_rows: int) -> pd.DataFrame:
    """Load one summary CSV and drop the first `warmup_rows` rows per config group."""
    df = pd.read_csv(path)

    # Sanity check: a single (max_num_seqs, random_input_len, random_output_len)
    # is expected per file, but we still group defensively.
    missing_cfg = [c for c in CONFIG_COLS if c not in df.columns]
    if missing_cfg:
        raise ValueError(f"{path} missing config columns: {missing_cfg}")

    n_groups = df.groupby(CONFIG_COLS, sort=False).ngroups
    if n_groups > 1:
        log.warning(
            "%s contains %d distinct (max_num_seqs, in_len, out_len) groups; "
            "warmup will be dropped per group",
            path, n_groups,
        )

    # Sort within each group by run_number (defensive — files appear sorted already)
    if "run_number" in df.columns:
        df = (
            df.sort_values(CONFIG_COLS + ["run_number"], kind="stable")
              .reset_index(drop=True)
        )

    return (
        df.groupby(CONFIG_COLS, sort=False, group_keys=False)
          .apply(lambda g: g.iloc[warmup_rows:])
          .reset_index(drop=True)
    )


# ---------------------------------------------------------------------------
# Aggregation
# ---------------------------------------------------------------------------


def harmonic_mean(values: np.ndarray) -> float:
    """Harmonic mean of strictly-positive values. Returns NaN on empty/invalid input."""
    arr = np.asarray(values, dtype=float)
    arr = arr[np.isfinite(arr) & (arr > 0)]
    if arr.size == 0:
        return float("nan")
    return float(arr.size / np.sum(1.0 / arr))


def arithmetic_mean(values: np.ndarray) -> float:
    """Arithmetic mean of finite values. Returns NaN on empty input."""
    arr = np.asarray(values, dtype=float)
    arr = arr[np.isfinite(arr)]
    if arr.size == 0:
        return float("nan")
    return float(np.mean(arr))


def metric_mean(values: np.ndarray, is_throughput: bool) -> float:
    """The right mean for the metric type.

    Throughput (rate) metrics → harmonic mean (correct for averaging rates).
    Latency (time) metrics → arithmetic mean.
    """
    return harmonic_mean(values) if is_throughput else arithmetic_mean(values)


def bootstrap_ratio_ci(
    baseline: np.ndarray,
    updated: np.ndarray,
    is_throughput: bool,
    n_iter: int = N_BOOTSTRAP,
    ci: float = CI_LEVEL,
) -> tuple[float, float, float]:
    """Bootstrap CI for the ratio updated_mean / baseline_mean.

    Returns (observed_ratio, lo, hi) where [lo, hi] is the central CI of the
    ratio. The mean used is `metric_mean` (harmonic for throughput, arithmetic
    for latency), matching the point estimate.

    Both arms are resampled independently with replacement. NaN/non-positive
    values are filtered out beforehand (harmonic mean is undefined on them).
    """
    base = np.asarray(baseline, dtype=float)
    upd = np.asarray(updated, dtype=float)
    base = base[np.isfinite(base) & (base > 0)]
    upd = upd[np.isfinite(upd) & (upd > 0)]

    if base.size == 0 or upd.size == 0:
        return float("nan"), float("nan"), float("nan")

    base_m = metric_mean(base, is_throughput)
    upd_m = metric_mean(upd, is_throughput)
    if not (np.isfinite(base_m) and np.isfinite(upd_m)) or base_m == 0:
        return float("nan"), float("nan"), float("nan")
    observed = upd_m / base_m

    # Vectorized bootstrap: sample n_iter * size index matrices, then aggregate.
    # This is ~100x faster than a Python loop and stays well under a second
    # even for 10k iterations on 20-element arrays.
    base_idx = _BOOT_RNG.integers(0, base.size, size=(n_iter, base.size))
    upd_idx = _BOOT_RNG.integers(0, upd.size, size=(n_iter, upd.size))
    base_samples = base[base_idx]
    upd_samples = upd[upd_idx]

    if is_throughput:
        # Harmonic mean per row: n / sum(1/x).
        base_means = base.size / np.sum(1.0 / base_samples, axis=1)
        upd_means = upd.size / np.sum(1.0 / upd_samples, axis=1)
    else:
        base_means = np.mean(base_samples, axis=1)
        upd_means = np.mean(upd_samples, axis=1)

    ratios = upd_means / base_means
    ratios = ratios[np.isfinite(ratios)]
    if ratios.size == 0:
        return observed, float("nan"), float("nan")

    low_p = (1.0 - ci) / 2.0 * 100.0
    high_p = (1.0 + ci) / 2.0 * 100.0
    lo, hi = np.percentile(ratios, [low_p, high_p])
    return float(observed), float(lo), float(hi)


def sample_std(values: np.ndarray) -> float:
    """Sample std (ddof=1) of finite values. Returns NaN if <2 points."""
    arr = np.asarray(values, dtype=float)
    arr = arr[np.isfinite(arr)]
    if arr.size < 2:
        return float("nan")
    return float(np.std(arr, ddof=1))


@dataclass
class MetricStat:
    """Aggregated stats for one (model, app, metric).

    `mean` is the arithmetic mean for latency metrics and harmonic mean for
    throughput metrics. `values` is the post-warmup raw array (kept for
    bootstrap CI computation against the baseline).
    """
    model: str
    app: str
    metric: str
    n: int
    mean: float
    std: float
    values: np.ndarray


def collect_stats(cfg: Config) -> tuple[list[MetricStat], dict[tuple[str, str], np.ndarray], str]:
    """
    Walk source_dir, load all CSVs, aggregate per (model, app, metric).
    Returns:
        per_app_stats: list of MetricStat (one per (model, app, metric))
        pooled_raw:    {(model, metric): pooled raw values across all non-baseline apps}
        baseline_app:  resolved baseline app name (case from inference)
    """
    discovered = discover_apps(cfg.source_dir)

    # Step 1: infer an app name for every subdirectory.
    # Multiple subdirs may map to the same app (e.g. one per model).
    dir_to_app: dict[str, str] = {}
    for raw_dir in discovered:
        app = infer_app_name(raw_dir, cfg.models)
        if app is None:
            log.warning(
                "Cannot infer app from directory %r (no configured model substring "
                "%s found in it); skipping",
                raw_dir, cfg.models,
            )
            continue
        dir_to_app[raw_dir] = app

    if not dir_to_app:
        raise RuntimeError(
            "No directories yielded an inferable app name. "
            "Check that subdirectory names contain a configured model name."
        )

    inferred_apps = set(dir_to_app.values())
    log.info("Directory → app inference:")
    for d, a in sorted(dir_to_app.items()):
        log.info("  %s → %s", d, a)

    baseline = resolve_baseline(inferred_apps, cfg.baseline_app)
    log.info("Inferred apps: %s (baseline=%s)", sorted(inferred_apps), baseline)

    # Step 2: aggregate. Group by inferred app — combine CSVs from multiple
    # directories that map to the same app.
    per_app: list[MetricStat] = []
    pooled: dict[tuple[str, str], list[np.ndarray]] = {}

    # raw[(app, model, metric)] -> list of arrays
    raw: dict[tuple[str, str, str], list[np.ndarray]] = {}

    for raw_dir, csv_entries in discovered.items():
        app = dir_to_app.get(raw_dir)
        if app is None:
            continue
        for path, _ in csv_entries:
            df = load_csv_post_warmup(path, cfg.warmup_rows)
            if df.empty:
                log.warning("%s has no rows after warmup, skipping", path)
                continue

            model_ids = df[MODEL_ID_COL].dropna().unique()
            if len(model_ids) != 1:
                log.warning(
                    "%s has %d distinct model_id values %s; skipping",
                    path, len(model_ids), list(model_ids),
                )
                continue
            model = map_model(str(model_ids[0]), cfg.models)
            if model is None:
                log.warning(
                    "model_id %r in %s does not match any configured model "
                    "(%s); skipping",
                    model_ids[0], path, cfg.models,
                )
                continue

            # Sanity: cross-check the model inferred from the directory name
            # against the model_id column. They should agree.
            dir_model_low = raw_dir.lower()
            if model.lower() not in dir_model_low:
                log.warning(
                    "Mismatch: directory %r suggests a different model than "
                    "%s (from model_id=%s in %s)",
                    raw_dir, model, model_ids[0], path,
                )

            for metric in cfg.metrics:
                if metric not in df.columns:
                    log.warning("%s missing metric column %s", path, metric)
                    continue
                vals = df[metric].to_numpy(dtype=float)
                raw.setdefault((app, model, metric), []).append(vals)

    for (app, model, metric), arrays in raw.items():
        combined = np.concatenate(arrays) if arrays else np.array([])
        is_throughput = _is_throughput_metric(metric)
        per_app.append(
            MetricStat(
                model=model,
                app=app,
                metric=metric,
                n=int(np.sum(np.isfinite(combined))),
                mean=metric_mean(combined, is_throughput),
                std=sample_std(combined),
                values=combined,
            )
        )
        if app.lower() != baseline.lower():
            pooled.setdefault((model, metric), []).append(combined)

    pooled_arrays = {k: np.concatenate(v) for k, v in pooled.items() if v}
    return per_app, pooled_arrays, baseline


# ---------------------------------------------------------------------------
# Shaping
# ---------------------------------------------------------------------------


def build_model_table(
    stats: list[MetricStat],
    pooled: dict[tuple[str, str], np.ndarray],
    model: str,
    metric: str,
    baseline_app: str,
    apps_order: list[str],
) -> pd.DataFrame:
    """
    One DataFrame per (model, metric) with rows for every app + the pooled aggregate.
    Columns: app, n, mean, std, pct_change, pct_lo, pct_hi.

    `pct_change` is signed % change vs baseline, with the convention that
    POSITIVE = REGRESSION for every metric:
      - latency-style metrics:    (app / baseline - 1) * 100        [arith mean]
      - throughput-style metrics: (1 - app / baseline) * 100        [harmonic mean]

    `pct_lo` / `pct_hi` are the 95% bootstrap CI bounds on `pct_change`,
    computed by independently resampling the app and baseline raw values and
    taking the percentile of the resulting ratio distribution. The interval is
    asymmetric in general; for the baseline row itself it is zero by definition.
    """
    rows = []
    by_app = {s.app: s for s in stats if s.model == model and s.metric == metric}

    base = by_app.get(baseline_app)
    base_mean = base.mean if base else float("nan")
    base_values = base.values if base is not None else np.array([])
    is_throughput = _is_throughput_metric(metric)

    def ratio_to_pct(r: float) -> float:
        if not np.isfinite(r):
            return float("nan")
        return (1.0 - r) * 100.0 if is_throughput else (r - 1.0) * 100.0

    def app_pct_and_ci(app_values: np.ndarray) -> tuple[float, float, float]:
        """Return (pct_change, pct_lo, pct_hi) for an app vs the baseline.

        For throughput metrics, the ratio CI [r_lo, r_hi] maps to the % bounds
        as (1 - r_hi)*100 .. (1 - r_lo)*100 — note the swap, because the
        transform is decreasing in r.
        """
        if base_values.size == 0:
            return float("nan"), float("nan"), float("nan")
        observed, r_lo, r_hi = bootstrap_ratio_ci(
            base_values, app_values, is_throughput
        )
        pct = ratio_to_pct(observed)
        if is_throughput:
            return pct, ratio_to_pct(r_hi), ratio_to_pct(r_lo)
        return pct, ratio_to_pct(r_lo), ratio_to_pct(r_hi)

    for app in apps_order:
        s = by_app.get(app)
        if s is None:
            continue
        is_baseline = app.lower() == baseline_app.lower()
        if is_baseline:
            pct_change, pct_lo, pct_hi = 0.0, 0.0, 0.0
        else:
            pct_change, pct_lo, pct_hi = app_pct_and_ci(s.values)
        rows.append(
            {
                "app": app,
                "is_baseline": is_baseline,
                "n": s.n,
                "mean": s.mean,
                "std": s.std,
                "pct_change": pct_change,
                "pct_lo": pct_lo,
                "pct_hi": pct_hi,
            }
        )

    pooled_arr = pooled.get((model, metric))
    if pooled_arr is not None and pooled_arr.size:
        pm = metric_mean(pooled_arr, is_throughput)
        ps = sample_std(pooled_arr)
        pct_change, pct_lo, pct_hi = app_pct_and_ci(pooled_arr)
        rows.append(
            {
                "app": COMBINED_LABEL,
                "is_baseline": False,
                "n": int(np.sum(np.isfinite(pooled_arr))),
                "mean": pm,
                "std": ps,
                "pct_change": pct_change,
                "pct_lo": pct_lo,
                "pct_hi": pct_hi,
            }
        )

    return pd.DataFrame(rows)


def _is_throughput_metric(metric: str) -> bool:
    """Higher-is-better metrics use the (1 - app/base) convention so positive = regression."""
    return "throughput" in metric.lower()


# ---------------------------------------------------------------------------
# XLSX
# ---------------------------------------------------------------------------


HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", start_color="305496")
BASELINE_FILL = PatternFill("solid", start_color="FFF2CC")
COMBINED_FILL = PatternFill("solid", start_color="E2EFDA")
DEFAULT_FONT = Font(name="Arial")
CENTER = Alignment(horizontal="center")


def _write_metric_sheet(wb: Workbook, sheet_name: str, df: pd.DataFrame, metric: str) -> None:
    ws = wb.create_sheet(sheet_name)
    ws.append([f"Metric: {metric}"])
    ws["A1"].font = Font(name="Arial", bold=True, size=12)
    ws.append([])

    headers = [
        "app", "n", "mean", "std",
        "% change vs baseline", "% CI low", "% CI high",
    ]
    ws.append(headers)
    header_row = ws.max_row
    for col, _ in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER

    for _, r in df.iterrows():
        ws.append(
            [
                r["app"],
                int(r["n"]) if np.isfinite(r["n"]) else None,
                round(float(r["mean"]), 4) if np.isfinite(r["mean"]) else None,
                round(float(r["std"]), 4) if np.isfinite(r["std"]) else None,
                round(float(r["pct_change"]), 2) if np.isfinite(r["pct_change"]) else None,
                round(float(r["pct_lo"]), 2) if np.isfinite(r["pct_lo"]) else None,
                round(float(r["pct_hi"]), 2) if np.isfinite(r["pct_hi"]) else None,
            ]
        )
        if r["is_baseline"]:
            for col in range(1, len(headers) + 1):
                ws.cell(row=ws.max_row, column=col).fill = BASELINE_FILL
        elif r["app"] == COMBINED_LABEL:
            for col in range(1, len(headers) + 1):
                ws.cell(row=ws.max_row, column=col).fill = COMBINED_FILL

    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 18 if col_idx == 1 else 14
        for row_idx in range(header_row, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.font is None or cell.font.name != "Arial":
                cell.font = DEFAULT_FONT

    _add_chart(ws, df, metric, headers)


def _add_chart(ws, df: pd.DataFrame, metric: str, headers: list[str]) -> None:
    """Embed a bar chart of % change vs baseline with asymmetric 95% CI error bars.

    The CI bounds in `pct_lo` / `pct_hi` are absolute % points. openpyxl
    ErrorBars need offsets from the bar height, so we write two helper columns
    (`% err minus`, `% err plus`) immediately after the headers and reference
    them as the error sources.
    """
    if df.empty:
        return
    n = len(df)
    header_row_idx = 3
    data_start = header_row_idx + 1
    data_end = data_start + n - 1
    pct_col = headers.index("% change vs baseline") + 1
    pct_lo_col = headers.index("% CI low") + 1
    pct_hi_col = headers.index("% CI high") + 1

    # Write helper columns for asymmetric error offsets (length-of-whisker, not bound).
    err_minus_col = len(headers) + 1
    err_plus_col = len(headers) + 2
    ws.cell(row=header_row_idx, column=err_minus_col, value="% err minus").font = HEADER_FONT
    ws.cell(row=header_row_idx, column=err_minus_col).fill = HEADER_FILL
    ws.cell(row=header_row_idx, column=err_plus_col, value="% err plus").font = HEADER_FONT
    ws.cell(row=header_row_idx, column=err_plus_col).fill = HEADER_FILL
    for i, (_, r) in enumerate(df.iterrows()):
        row_idx = data_start + i
        pc = float(r["pct_change"]) if np.isfinite(r["pct_change"]) else None
        lo = float(r["pct_lo"]) if np.isfinite(r["pct_lo"]) else None
        hi = float(r["pct_hi"]) if np.isfinite(r["pct_hi"]) else None
        minus = round(pc - lo, 4) if pc is not None and lo is not None else None
        plus = round(hi - pc, 4) if pc is not None and hi is not None else None
        # Clamp tiny negatives from rounding (shouldn't happen but be safe).
        if minus is not None and minus < 0:
            minus = 0.0
        if plus is not None and plus < 0:
            plus = 0.0
        ws.cell(row=row_idx, column=err_minus_col, value=minus)
        ws.cell(row=row_idx, column=err_plus_col, value=plus)

    chart = BarChart()
    chart.type = "col"
    chart.style = 2
    chart.title = f"{metric} — % change vs baseline (positive = regression)"
    chart.y_axis.title = "% change vs baseline (0 = baseline)"
    chart.x_axis.title = "App"
    chart.legend = None

    data_ref = Reference(
        ws, min_col=pct_col, max_col=pct_col,
        min_row=header_row_idx, max_row=data_end,
    )
    chart.add_data(data_ref, titles_from_data=True)

    cats = Reference(ws, min_col=1, min_row=data_start, max_row=data_end)
    chart.set_categories(cats)

    # Asymmetric error bars: separate plus and minus refs.
    plus_range = (
        f"'{ws.title}'!${get_column_letter(err_plus_col)}${data_start}"
        f":${get_column_letter(err_plus_col)}${data_end}"
    )
    minus_range = (
        f"'{ws.title}'!${get_column_letter(err_minus_col)}${data_start}"
        f":${get_column_letter(err_minus_col)}${data_end}"
    )
    err_bars = ErrorBars(
        errDir="y", errBarType="both", errValType="cust",
        plus=NumDataSource(numRef=NumRef(f=plus_range)),
        minus=NumDataSource(numRef=NumRef(f=minus_range)),
    )
    chart.series[0].errBars = err_bars

    chart.height = 10
    chart.width = 20
    ws.add_chart(chart, f"H{header_row_idx}")


def write_model_workbook(
    out_path: Path,
    model: str,
    metric_tables: dict[str, pd.DataFrame],
) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    summary = wb.create_sheet("summary")
    summary.append([f"Model: {model}"])
    summary["A1"].font = Font(name="Arial", bold=True, size=14)
    summary.append([])
    summary.append([
        "Metric", "App", "n", "mean", "std",
        "% change vs baseline", "% CI low", "% CI high",
    ])
    n_summary_cols = 8
    for col_idx in range(1, n_summary_cols + 1):
        cell = summary.cell(row=3, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER

    for metric, df in metric_tables.items():
        for _, r in df.iterrows():
            summary.append(
                [
                    metric,
                    r["app"],
                    int(r["n"]) if np.isfinite(r["n"]) else None,
                    round(float(r["mean"]), 4) if np.isfinite(r["mean"]) else None,
                    round(float(r["std"]), 4) if np.isfinite(r["std"]) else None,
                    round(float(r["pct_change"]), 2) if np.isfinite(r["pct_change"]) else None,
                    round(float(r["pct_lo"]), 2) if np.isfinite(r["pct_lo"]) else None,
                    round(float(r["pct_hi"]), 2) if np.isfinite(r["pct_hi"]) else None,
                ]
            )
            if r["is_baseline"]:
                for c in range(1, n_summary_cols + 1):
                    summary.cell(row=summary.max_row, column=c).fill = BASELINE_FILL
            elif r["app"] == COMBINED_LABEL:
                for c in range(1, n_summary_cols + 1):
                    summary.cell(row=summary.max_row, column=c).fill = COMBINED_FILL

    summary.column_dimensions["A"].width = 26
    summary.column_dimensions["B"].width = 22
    for col in "CDEFGH":
        summary.column_dimensions[col].width = 14

    for metric, df in metric_tables.items():
        # Excel sheet names are capped at 31 chars and disallow some chars.
        safe = metric.replace("/", "_")[:31]
        _write_metric_sheet(wb, safe, df, metric)

    wb.save(out_path)


# ---------------------------------------------------------------------------
# Charts (matplotlib)
# ---------------------------------------------------------------------------


def _scientific_style(presentation: ChartPresentation) -> None:
    plt.rcParams.update(
        {
            "font.family": "DejaVu Sans",
            "font.size": presentation.font_size("base"),
            "axes.titlesize": presentation.font_size("title"),
            "axes.labelsize": presentation.font_size("axes_label"),
            "xtick.labelsize": presentation.font_size("xtick"),
            "ytick.labelsize": presentation.font_size("ytick"),
            "axes.spines.top": False,
            "axes.spines.right": False,
            "axes.grid": True,
            "grid.alpha": 0.3,
            "grid.linestyle": "--",
            "figure.dpi": 150,
        }
    )


_TAB10 = [
    "#1f77b4", "#ff7f0e", "#2ca02c", "#9467bd", "#8c564b",
    "#e377c2", "#17becf", "#bcbd22",
]
_BASELINE_COLOR = "#7f7f7f"


def _presentation_df(df: pd.DataFrame, presentation: ChartPresentation) -> pd.DataFrame:
    if df.empty:
        return df
    return df[~df["app"].apply(presentation.hide_app)].reset_index(drop=True)


def _build_color_map(
    apps: list[str],
    baseline_app: str,
    presentation: ChartPresentation,
) -> dict[str, str]:
    """Stable color per app: tab10 palette assigned in sorted order.
    Baseline overrides to grey; special aggregates use presentation overrides.
    """
    palette_apps = sorted(
        (a for a in apps if a.lower() != baseline_app.lower()),
        key=str.lower,
    )
    mapping = {a: _TAB10[i % len(_TAB10)] for i, a in enumerate(palette_apps)}
    mapping[baseline_app] = presentation.color(baseline_app, _BASELINE_COLOR)
    return mapping


def _bar_color(app: str, color_map: dict[str, str], presentation: ChartPresentation) -> str:
    if app in DEFAULT_SPECIAL_COLORS:
        return presentation.color(app, DEFAULT_SPECIAL_COLORS[app])
    return presentation.color(app, color_map.get(app, _TAB10[0]))


def _build_hatch_map(
    apps: list[str],
    baseline_app: str,
    presentation: ChartPresentation,
) -> dict[str, str]:
    patterns = presentation.hatch_patterns()
    palette_apps = sorted(
        (a for a in apps if a.lower() != baseline_app.lower()),
        key=str.lower,
    )
    mapping = {a: patterns[i % len(patterns)] for i, a in enumerate(palette_apps)}
    mapping[baseline_app] = presentation.explicit_hatch(baseline_app) or ""
    for key in DEFAULT_SPECIAL_COLORS:
        explicit = presentation.explicit_hatch(key)
        if explicit is not None:
            mapping[key] = explicit
    for app in apps:
        explicit = presentation.explicit_hatch(app)
        if explicit is not None:
            mapping[app] = explicit
    return mapping


def _bar_hatch(app: str, hatch_map: dict[str, str], presentation: ChartPresentation) -> str:
    explicit = presentation.explicit_hatch(app)
    if explicit is not None:
        return explicit
    return hatch_map.get(app, "")


def _format_pct_label(v: float) -> str:
    """Signed % label: +5.2%, -20.1%, 0.0%."""
    if not np.isfinite(v):
        return ""
    if abs(v) < 0.05:
        return "0.0%"
    return f"{v:+.1f}%"


def _annotate_bars(ax, bars, values, errors, label_keys, presentation: ChartPresentation) -> None:
    """Place a signed % label above (positive bars) or below (negative bars) each bar.

    Errors are added so labels clear the error caps. Labels are skipped for NaN bars.
    """
    if len(bars) == 0:
        return
    finite = np.asarray(values)[np.isfinite(values)]
    span = float(np.ptp(finite)) if finite.size else 1.0
    pad = max(span * 0.025, 0.5)  # in data units (% points)
    for bar, v, e, label_key in zip(bars, values, errors, label_keys):
        if not np.isfinite(v):
            continue
        e = e if np.isfinite(e) else 0.0
        x = bar.get_x() + bar.get_width() / 2
        overflow_top = max(v, v + e)
        if v >= 0 and overflow_top >= CHART_YMAX - CHART_OVERFLOW_EPS:
            ax.annotate(
                _format_pct_label(v),
                xy=(x, CHART_YMAX - 0.02),
                xytext=(x + bar.get_width() * 0.7, CHART_YMAX - pad * 0.5),
                ha="left",
                va="top",
                fontsize=presentation.font_size("annotation", label_key),
                fontweight="bold",
                color="black",
                clip_on=False,
                bbox={
                    "boxstyle": "round,pad=0.22",
                    "facecolor": "white",
                    "edgecolor": "black",
                    "linewidth": 0.8,
                    "alpha": 0.96,
                },
                arrowprops={
                    "arrowstyle": "-|>",
                    "color": "black",
                    "linewidth": 0.9,
                    "shrinkA": 4,
                    "shrinkB": 4,
                    "mutation_scale": 8,
                },
                path_effects=[pe.withStroke(linewidth=1.0, foreground="white")],
            )
            continue
        if v >= 0:
            y = min(v + e + pad, CHART_YMAX + pad * 0.35)
            va = "bottom"
        else:
            y = v - e - pad
            va = "top"
        ax.text(
            x,
            y,
            _format_pct_label(v),
            ha="center",
            va=va,
            fontsize=presentation.font_size("annotation", label_key),
            fontweight="bold",
            color="black",
            clip_on=False,
            bbox={
                "boxstyle": "round,pad=0.18",
                "facecolor": "white",
                "edgecolor": "none",
                "alpha": 0.85,
            },
            path_effects=[pe.withStroke(linewidth=1.0, foreground="white")],
        )


def _draw_broken_bar_markers(ax, bars, values, errors) -> None:
    """Draw a bar-break marker where a positive bar or its upper whisker crosses the fixed y-limit."""
    for bar, v, e in zip(bars, values, errors):
        if not np.isfinite(v):
            continue
        e = e if np.isfinite(e) else 0.0
        overflow_top = max(v, v + e)
        if v < 0 or overflow_top < CHART_YMAX - CHART_OVERFLOW_EPS:
            continue

        x0 = bar.get_x()
        width = bar.get_width()
        y = CHART_YMAX - 0.15
        dx = width * 0.18
        gap = width * 0.12
        ax.plot(
            [x0 + width * 0.5 - gap - dx, x0 + width * 0.5 - gap + dx],
            [y - 0.28, y + 0.28],
            color="black",
            linewidth=1.2,
            clip_on=False,
            zorder=6,
        )
        ax.plot(
            [x0 + width * 0.5 + gap - dx, x0 + width * 0.5 + gap + dx],
            [y - 0.28, y + 0.28],
            color="black",
            linewidth=1.2,
            clip_on=False,
            zorder=6,
        )


def _expand_ylim(ax, *value_arrays) -> None:
    """Use a fixed top limit and only expand the lower bound as needed."""
    all_vals: list[float] = []
    for arr in value_arrays:
        a = np.asarray(arr, dtype=float)
        all_vals.extend(a[np.isfinite(a)].tolist())
    lo = min(all_vals) if all_vals else 0.0
    lo = min(lo, 0.0)
    lower_span = max(CHART_YMAX - lo, 1.0)
    lower_pad = max(lower_span * 0.08, 0.5)
    ax.set_ylim(lo - lower_pad, CHART_YMAX)


def plot_per_model(
    df: pd.DataFrame,
    model: str,
    metric: str,
    color_map: dict[str, str],
    hatch_map: dict[str, str],
    presentation: ChartPresentation,
    out_path: Path,
) -> None:
    df = _presentation_df(df, presentation)
    if df.empty:
        return
    _scientific_style(presentation)
    apps = df["app"].tolist()
    pcts = df["pct_change"].to_numpy(dtype=float)
    los = df["pct_lo"].to_numpy(dtype=float)
    his = df["pct_hi"].to_numpy(dtype=float)
    # Asymmetric error: distance from bar height to each CI bound.
    err_minus = np.where(np.isfinite(pcts) & np.isfinite(los), pcts - los, 0.0)
    err_plus = np.where(np.isfinite(pcts) & np.isfinite(his), his - pcts, 0.0)
    err_minus = np.clip(err_minus, 0.0, None)
    err_plus = np.clip(err_plus, 0.0, None)
    errs = np.vstack([err_minus, err_plus])

    colors = [_bar_color(a, color_map, presentation) for a in apps]
    hatches = [_bar_hatch(a, hatch_map, presentation) for a in apps]

    fig, ax = plt.subplots(figsize=(max(6, 1.1 * len(apps) + 2), 5.5))
    ax.yaxis.grid(True, linestyle="--", alpha=0.3)
    ax.xaxis.grid(False)
    x = np.arange(len(apps))
    bars = ax.bar(
        x, pcts, yerr=errs, capsize=4, color=colors,
        edgecolor="black", linewidth=0.6,
    )
    for bar, hatch in zip(bars, hatches):
        bar.set_hatch(hatch)
    ax.axhline(0.0, color="black", linewidth=0.8, linestyle=":", alpha=0.7)
    ax.set_xticks(x)
    ax.set_xticklabels([presentation.display_label(app) for app in apps], rotation=30, ha="right")
    for tick, app in zip(ax.get_xticklabels(), apps):
        tick.set_fontsize(presentation.font_size("xtick", app))
    ax.set_ylabel(_metric_axis_label(metric, presentation))
    # Annotation pad uses the larger of the two whiskers per bar.
    ann_errs = np.maximum(err_minus, err_plus)
    _annotate_bars(ax, bars, pcts, ann_errs, apps, presentation)
    _draw_broken_bar_markers(ax, bars, pcts, err_plus)
    _expand_ylim(ax, pcts - err_minus, pcts + err_plus)
    fig.tight_layout()
    fig.savefig(out_path, dpi=300, bbox_inches="tight")
    plt.close(fig)


def _legend_grid(n_entries: int, max_rows: int = 2) -> tuple[int, int]:
    """Columns/rows for a wrapped legend spread across at most `max_rows` rows."""
    cols = max(1, math.ceil(n_entries / max_rows))
    rows = math.ceil(n_entries / cols)
    return cols, rows


def plot_combined(
    per_model_df: dict[str, pd.DataFrame],
    metric: str,
    color_map: dict[str, str],
    hatch_map: dict[str, str],
    presentation: ChartPresentation,
    out_path: Path,
) -> None:
    """One chart, one metric, models as groups along the x-axis.

    A final group on the right, labeled `across_models`, contains a single bar:
    the harmonic mean (across models) of each model's `all_apps_combined` value.
    No error bar on that bar — by design.
    """
    models = [m for m, df in per_model_df.items() if not df.empty]
    if len(models) < 2:
        return

    _scientific_style(presentation)

    filtered_per_model_df = {
        model: _presentation_df(df, presentation)
        for model, df in per_model_df.items()
    }
    models = [m for m, df in filtered_per_model_df.items() if not df.empty]
    if len(models) < 2:
        return

    apps_per_model = {m: filtered_per_model_df[m]["app"].tolist() for m in models}
    union_apps = []
    seen = set()
    for m in models:
        for a in apps_per_model[m]:
            if a not in seen:
                seen.add(a)
                union_apps.append(a)
    # Order by the configured app_order so apps present in only some models
    # (e.g. the mixtral-only MoE app) sit with the other App-N entries; keep the
    # COMBINED_LABEL aggregate pinned last.
    order_key = _order_key(presentation.app_order)
    union_apps.sort(key=lambda a: (a == COMBINED_LABEL, order_key(a)))

    # The across_models aggregate is the harmonic mean (across models) of each
    # model's all_apps_combined RATIO. Since pct_change is a linear function of
    # the ratio (1 - r or r - 1), we recover the ratio per model, harmonic-mean
    # those ratios, then convert back to pct_change.
    is_throughput = _is_throughput_metric(metric)

    def pct_to_ratio(p: float) -> float:
        return 1.0 - p / 100.0 if is_throughput else 1.0 + p / 100.0

    def ratio_to_pct(r: float) -> float:
        return (1.0 - r) * 100.0 if is_throughput else (r - 1.0) * 100.0

    pooled_ratios: list[float] = []
    for m in models:
        df = filtered_per_model_df[m]
        row = df[df["app"] == COMBINED_LABEL]
        if not row.empty:
            v = float(row["pct_change"].iloc[0])
            if np.isfinite(v):
                r = pct_to_ratio(v)
                if r > 0:
                    pooled_ratios.append(r)
    across_models_value = (
        ratio_to_pct(harmonic_mean(np.array(pooled_ratios)))
        if pooled_ratios else float("nan")
    )

    n_groups = len(models) + 1  # +1 for the across_models group
    n_bars = len(union_apps)
    bar_w = 0.9 / n_bars
    x = np.arange(n_groups, dtype=float)
    if n_groups >= 2:
        x[-1] = x[-2] + 0.72

    # Wrap the legend across up to `max_cols` columns so big text doesn't force
    # one ultra-wide row, and grow the figure to host the legend band rather than
    # shrinking the axes — keeps the plot full-height (paper-readable).
    legend_entries = n_bars + (1 if np.isfinite(across_models_value) else 0)
    legend_cols, legend_rows = _legend_grid(legend_entries)

    PLOT_HEIGHT_IN = 5.5  # desired drawable plot height, kept constant
    legend_fs = presentation.font_size("legend")
    legend_band_in = legend_rows * (legend_fs / 72.0 * 1.7) + 0.35
    x_label_margin_in = 1.0
    fig_width = max(10.5, 2.35 * n_groups + 0.85 * n_bars)
    fig_height = PLOT_HEIGHT_IN + legend_band_in + x_label_margin_in
    fig, ax = plt.subplots(figsize=(fig_width, fig_height))
    ax.yaxis.grid(True, linestyle="--", alpha=0.3)
    ax.xaxis.grid(False)

    # Track every bar height + error so we can size the y-axis at the end.
    all_low_edges: list[float] = []
    all_high_edges: list[float] = []

    for i, app in enumerate(union_apps):
        heights = []
        err_minus = []
        err_plus = []
        for m in models:
            df = filtered_per_model_df[m]
            row = df[df["app"] == app]
            if row.empty:
                heights.append(np.nan)
                err_minus.append(0.0)
                err_plus.append(0.0)
            else:
                v = float(row["pct_change"].iloc[0])
                lo = float(row["pct_lo"].iloc[0])
                hi = float(row["pct_hi"].iloc[0])
                heights.append(v if np.isfinite(v) else np.nan)
                em = (v - lo) if (np.isfinite(v) and np.isfinite(lo)) else 0.0
                ep = (hi - v) if (np.isfinite(v) and np.isfinite(hi)) else 0.0
                err_minus.append(max(em, 0.0))
                err_plus.append(max(ep, 0.0))
        # No bar for any individual app in the across_models group.
        heights.append(np.nan)
        err_minus.append(0.0)
        err_plus.append(0.0)

        offsets = x - 0.45 + bar_w * (i + 0.5)
        color = _bar_color(app, color_map, presentation)
        bars = ax.bar(
            offsets, heights, bar_w,
            yerr=np.vstack([err_minus, err_plus]),
            capsize=3,
            label=presentation.series_display_name(app),
            color=color,
            edgecolor="black",
            linewidth=0.5,
        )
        hatch = _bar_hatch(app, hatch_map, presentation)
        for bar in bars:
            bar.set_hatch(hatch)
        ann_errs = [max(a, b) for a, b in zip(err_minus, err_plus)]
        _annotate_bars(ax, bars, heights, ann_errs, [app] * len(heights), presentation)
        _draw_broken_bar_markers(ax, bars, heights, err_plus)
        # Track the actual extents (bar top + plus, bar bottom - minus).
        for h, em, ep in zip(heights, err_minus, err_plus):
            if np.isfinite(h):
                all_low_edges.append(h - em)
                all_high_edges.append(h + ep)

    if np.isfinite(across_models_value):
        am_bars = ax.bar(
            x[-1], across_models_value, bar_w * 1.5,
            color=_bar_color(ACROSS_MODELS_LABEL, color_map, presentation),
            edgecolor="black",
            linewidth=0.5,
            label=presentation.series_display_name(ACROSS_MODELS_LABEL),
        )
        for bar in am_bars:
            bar.set_hatch(_bar_hatch(ACROSS_MODELS_LABEL, hatch_map, presentation))
        _annotate_bars(
            ax,
            am_bars,
            [across_models_value],
            [0.0],
            [ACROSS_MODELS_LABEL],
            presentation,
        )
        _draw_broken_bar_markers(ax, am_bars, [across_models_value], [0.0])
        all_low_edges.append(across_models_value)
        all_high_edges.append(across_models_value)

    ax.axhline(0.0, color="black", linewidth=0.8, linestyle=":", alpha=0.7)
    ax.set_xticks(x)
    combined_tick_keys = models + [ACROSS_MODELS_LABEL]
    across_models_tick_label = presentation.display_label(ACROSS_MODELS_LABEL).replace("-", "-\n")
    combined_tick_labels = models + [across_models_tick_label]
    ax.set_xticklabels(combined_tick_labels)
    for tick, key in zip(ax.get_xticklabels(), combined_tick_keys):
        tick.set_fontsize(presentation.font_size("xtick", key))
        tick.set_fontweight(presentation.font_weight("xtick", key))
    ax.set_ylabel(_metric_axis_label(metric, presentation))
    legend = ax.legend(
        loc="lower center",
        bbox_to_anchor=(0.5, 1.02),
        fontsize=presentation.font_size("legend"),
        frameon=False,
        ncol=legend_cols,
        columnspacing=0.9,
        handlelength=1.4,
        handletextpad=0.5,
        borderaxespad=0.0,
    )
    if legend is not None:
        legend_keys = union_apps + ([ACROSS_MODELS_LABEL] if np.isfinite(across_models_value) else [])
        for text, key in zip(legend.get_texts(), legend_keys):
            text.set_fontsize(presentation.font_size("legend", key))
            text.set_fontweight(presentation.font_weight("legend", key))
    _expand_ylim(ax, np.array(all_low_edges), np.array(all_high_edges))
    # Reserve the top legend band (computed in inches above) so the axes keep
    # their full height; the grown figure absorbs the legend.
    top_rect = 1.0 - legend_band_in / fig_height
    fig.tight_layout(rect=(0.0, 0.0, 1.0, top_rect))
    fig.savefig(out_path, dpi=300, bbox_inches="tight")
    plt.close(fig)


# ---------------------------------------------------------------------------
# Run orchestration
# ---------------------------------------------------------------------------


def prepare_output_dir(output_dir: Path) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    (output_dir / "xlsx").mkdir(parents=True, exist_ok=True)
    (output_dir / "charts" / "png").mkdir(parents=True, exist_ok=True)
    return output_dir


def _order_key(app_order: list[str]):
    """Sort key putting apps listed in `app_order` first (in that order), then
    any unlisted apps alphabetically after them."""
    index = {a: i for i, a in enumerate(app_order)}
    return lambda a: (index.get(a, len(app_order)), a.lower())


def _ordered_apps(stats: list[MetricStat], baseline_app: str, app_order: list[str]) -> list[str]:
    """Stable order: baseline first, then other apps by the configured
    `app_order` (unlisted apps follow alphabetically)."""
    apps = sorted({s.app for s in stats}, key=_order_key(app_order))
    others = [a for a in apps if a.lower() != baseline_app.lower()]
    return [baseline_app] + others


def run(cfg: Config) -> Path:
    if not cfg.source_dir.is_dir():
        raise RuntimeError(f"source_dir does not exist: {cfg.source_dir}")

    run_dir = prepare_output_dir(cfg.output_dir)
    log.info("Output directory: %s", run_dir)

    shutil.copy(cfg_path_global, run_dir / "config_used.json")

    stats, pooled, baseline = collect_stats(cfg)

    # The baseline is 0% by definition, so hide its bar from the charts.
    cfg.chart_presentation.hidden_apps.add(baseline)

    discovered_models = sorted({s.model for s in stats})
    if not discovered_models:
        raise RuntimeError("No CSVs matched any configured model. Nothing to do.")
    log.info("Models with data: %s", discovered_models)

    apps_order = _ordered_apps(stats, baseline, cfg.chart_presentation.app_order)
    color_map = _build_color_map(apps_order, baseline, cfg.chart_presentation)
    hatch_map = _build_hatch_map(apps_order + [COMBINED_LABEL, ACROSS_MODELS_LABEL], baseline, cfg.chart_presentation)

    per_metric_per_model: dict[str, dict[str, pd.DataFrame]] = {}
    for model in discovered_models:
        per_metric_per_model[model] = {}
        for metric in cfg.metrics:
            df = build_model_table(stats, pooled, model, metric, baseline, apps_order)
            per_metric_per_model[model][metric] = df

    for model, metric_tables in per_metric_per_model.items():
        xlsx_path = run_dir / "xlsx" / f"{model}.xlsx"
        write_model_workbook(xlsx_path, model, metric_tables)
        log.info("Wrote %s", xlsx_path)

        for metric, df in metric_tables.items():
            png_path = run_dir / "charts" / "png" / f"{model}_{metric}.png"
            plot_per_model(
                df,
                model,
                metric,
                color_map,
                hatch_map,
                cfg.chart_presentation,
                png_path,
            )
            log.info("Wrote %s", png_path)

    if len(discovered_models) >= 2:
        for metric in cfg.metrics:
            per_model_df = {m: per_metric_per_model[m][metric] for m in discovered_models}
            png_path = run_dir / "charts" / "png" / f"combined_{metric}.png"
            plot_combined(
                per_model_df,
                metric,
                color_map,
                hatch_map,
                cfg.chart_presentation,
                png_path,
            )
            log.info("Wrote %s", png_path)
    else:
        log.info("Only one model with data; skipping combined charts.")

    return run_dir


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


cfg_path_global: Path = Path(".")  # set in main()


def main(argv: Iterable[str] | None = None) -> int:
    global cfg_path_global

    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--config", required=True, type=Path, help="Path to JSON config")
    parser.add_argument("-v", "--verbose", action="count", default=0)
    args = parser.parse_args(list(argv) if argv is not None else None)

    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(levelname)s %(name)s | %(message)s",
    )

    cfg_path_global = args.config
    cfg = Config.from_json(args.config)
    out = run(cfg)
    print(f"Run complete: {out}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
